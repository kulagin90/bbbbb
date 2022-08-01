from openpyxl import load_workbook
from aiogram import types, Bot
from loader import dp


@dp.message_handler(text='/start')
async def command_start(message: types.Message):
    await message.answer(
        f'Привет {message.from_user.full_name}, я бот, который поможет найти машину на' + 'https://cars.av.by/')
    await message.answer('для поиска автомобиля введи : poisk')


@dp.message_handler(text='poisk')
async def pk(message):
    wb = load_workbook('C:/Users/kulag/PycharmProjects/bbbbb/Bot_aiogram-master/handlers/users/auto.xlsx')
    wb.active = 0
    sheet = wb.active
    for i in range(1, sheet.max_row + 1):
        val = sheet['A' + str(i)].value #, sheet['B' + str(i)].value# sheet['C' + str(i)].value,sheet['D' + str(i)].value, sheet['E' + str(i)].value, sheet['F' + str(i)].value
        print(val)
        await message.answer(f'{val}')
